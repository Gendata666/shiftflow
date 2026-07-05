"""
Excel renderer — reproduces the client-approved layout of
beach_bar_grafik.xlsx generically for any spec: title bar, week blocks with
day names + date numbers, per-shift colors from the catalog, per-duration
summary columns, legend, quota line, plus a second sheet with the rule-by-rule
verification report.
"""

from __future__ import annotations

from collections import Counter
from datetime import timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.spec_draft import min_to_hhmm
from packages.scheduler.spec import OFF, Finding, GenerateReport, ScheduleSpec
from packages.scheduler.verifier import build_grid

DAY_NAMES_BG = ["Пон", "Вт", "Ср", "Чет", "Пет", "Съб", "Нед"]

NAVY = "1A2E44"
STEEL = "2C4A63"
ROW_BG = "E8EEF5"
OFF_BG = "D9D9D9"
WHITE = "FFFFFF"

_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
_thin = Side(style="thin", color="B0B0B0")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(hex_color: str) -> PatternFill:
    color = hex_color.lstrip("#").upper()
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def render_xlsx(spec: ScheduleSpec, report: GenerateReport, month_label: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Месечен График"
    grid = build_grid(report.result.assignments)
    days = spec.days()
    num_weeks = spec.num_weeks()
    durations = sorted({s.duration_h for s in spec.shifts})
    n_day_cols = len(days)
    first_sum_col = 2 + n_day_cols

    # Title bar
    last_col = first_sum_col + len(durations) - 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = ws.cell(row=1, column=1)
    title.value = f"МЕСЕЧЕН РАБОТЕН ГРАФИК — {spec.venue_name or 'ShiftFlow'}"
    if month_label:
        title.value += f" — {month_label}"
    title.font = Font(bold=True, size=13, color=WHITE)
    title.fill = _fill(NAVY)
    title.alignment = _center
    ws.row_dimensions[1].height = 26

    # Header: staff col + week blocks + duration summary columns
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
    hdr = ws.cell(row=2, column=1, value="Служител")
    hdr.font = Font(bold=True, size=10, color=WHITE)
    hdr.fill = _fill(NAVY)
    hdr.alignment = _center
    ws.column_dimensions["A"].width = 12

    for w in range(num_weeks):
        c0 = 2 + w * 7
        c1 = min(c0 + 6, 1 + n_day_cols)
        ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c1)
        cell = ws.cell(row=2, column=c0, value=f"СЕДМИЦА  {w + 1}")
        cell.font = Font(bold=True, size=9, color=WHITE)
        cell.fill = _fill(STEEL)
        cell.alignment = _center

    for i, day in enumerate(days):
        col = 2 + i
        ws.column_dimensions[get_column_letter(col)].width = 8
        name = ws.cell(row=3, column=col, value=DAY_NAMES_BG[day.weekday()])
        name.font = Font(bold=True, size=8, color=WHITE)
        name.fill = _fill(STEEL)
        name.alignment = _center
        num = ws.cell(row=4, column=col, value=day.day)
        num.font = Font(size=7)
        num.alignment = _center

    duration_color = {}
    for dur in durations:
        for s in spec.shifts:
            if s.duration_h == dur:
                duration_color[dur] = s.color_hex
                break
    for j, dur in enumerate(durations):
        col = first_sum_col + j
        ws.merge_cells(start_row=2, start_column=col, end_row=4, end_column=col)
        label = f"{dur:g}ч\nсмени"
        cell = ws.cell(row=2, column=col, value=label)
        cell.font = Font(bold=True, size=8)
        cell.fill = _fill(duration_color[dur])
        cell.alignment = _center
        ws.column_dimensions[get_column_letter(col)].width = 7

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 13

    # Staff rows
    row = 5
    for member in spec.staff:
        ws.row_dimensions[row].height = 24
        name_cell = ws.cell(row=row, column=1, value=member.name)
        name_cell.font = Font(bold=True, size=10)
        name_cell.fill = _fill(ROW_BG)
        name_cell.alignment = Alignment(vertical="center")
        name_cell.border = _border

        week_counts: Counter = Counter()
        for i, day in enumerate(days):
            cell = ws.cell(row=row, column=2 + i)
            shift_id = grid.get((member.id, day))
            cell.alignment = _center
            cell.border = _border
            cell.font = Font(size=7.5)
            if shift_id in (None, OFF):
                cell.value = "почивка"
                cell.fill = _fill(OFF_BG)
            else:
                shift = spec.shift_by_id(shift_id)
                cell.value = f"{min_to_hhmm(shift.start_min)}–{min_to_hhmm(shift.end_min)}"
                cell.fill = _fill(shift.color_hex)
                week_counts[shift.duration_h] += 1

        for j, dur in enumerate(durations):
            cell = ws.cell(row=row, column=first_sum_col + j, value=week_counts.get(dur, 0))
            cell.font = Font(bold=True, size=11)
            cell.fill = _fill(duration_color[dur])
            cell.alignment = _center
            cell.border = _border
        row += 1

    # Legend
    row += 1
    legend_row = row
    ws.cell(row=legend_row, column=1, value="ЛЕГЕНДА:").font = Font(bold=True, size=8)
    col = 3
    for s in spec.shifts:
        ws.merge_cells(start_row=legend_row, start_column=col, end_row=legend_row, end_column=col + 1)
        cell = ws.cell(row=legend_row, column=col)
        cell.value = f"{s.duration_h:g}ч  {min_to_hhmm(s.start_min)}–{min_to_hhmm(s.end_min)}"
        cell.font = Font(size=7.5)
        cell.fill = _fill(s.color_hex)
        cell.alignment = _center
        col += 2

    # Verification line
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    ok = report.clean
    relaxed = report.result.relaxed_rule_ids
    if ok and not relaxed:
        note = "✓ Всички правила са спазени (автоматична проверка)"
    elif relaxed:
        note = f"⚠ Невъзможни за пълно спазване правила: {', '.join(relaxed)} — виж лист „Проверка“"
    else:
        note = "⚠ Има нарушени правила — виж лист „Проверка“"
    note_cell = ws.cell(row=row, column=1, value=note)
    note_cell.font = Font(size=7.5, bold=not ok)

    _append_verification_sheet(wb, spec, report)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _append_verification_sheet(wb: Workbook, spec: ScheduleSpec, report: GenerateReport) -> None:
    ws = wb.create_sheet("Проверка")
    headers = ["Правило", "Описание", "Статус", "Нарушения", "Детайли"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, size=9, color=WHITE)
        cell.fill = _fill(NAVY)
    widths = [22, 40, 12, 10, 60]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    status_bg = {"ok": "C6EFCE", "relaxed": "FFEB9C", "violated": "FFC7CE"}
    status_label = {"ok": "OK", "relaxed": "невъзможно — минимизирано", "violated": "нарушено"}
    for r, f in enumerate(report.findings, start=2):
        rule = spec.rule_by_id(f.rule_id)
        ws.cell(row=r, column=1, value=f.rule_id).font = Font(size=8)
        ws.cell(row=r, column=2, value=rule.description).font = Font(size=8)
        st = ws.cell(row=r, column=3, value=status_label[f.status])
        st.font = Font(size=8, bold=f.status != "ok")
        st.fill = _fill(status_bg[f.status])
        ws.cell(row=r, column=4, value=f.violations).font = Font(size=8)
        ws.cell(row=r, column=5, value=f.message_bg if f.status != "ok" else "").font = Font(size=8)
