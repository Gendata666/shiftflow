from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import io

from app.core.database import get_db
from app.core.deps import require_manager
from app.models.user import User, StaffProfile
from app.models.schedule import Schedule, ShiftAssignment, SchedulePeriod
from app.models.shift_type import ShiftType

router = APIRouter()


async def _load_schedule_data(schedule_id: str, tenant_id: str, db: AsyncSession):
    sched_res = await db.execute(
        select(Schedule).where(Schedule.id == schedule_id, Schedule.tenant_id == tenant_id)
    )
    sched = sched_res.scalar_one_or_none()
    if not sched:
        raise HTTPException(status_code=404, detail="Schedule not found")

    period_res = await db.execute(select(SchedulePeriod).where(SchedulePeriod.id == sched.period_id))
    period = period_res.scalar_one()

    assign_res = await db.execute(
        select(ShiftAssignment, StaffProfile, User, ShiftType)
        .join(StaffProfile, StaffProfile.id == ShiftAssignment.staff_id)
        .join(User, User.id == StaffProfile.user_id)
        .join(ShiftType, ShiftType.id == ShiftAssignment.shift_type_id)
        .where(ShiftAssignment.schedule_id == schedule_id)
        .order_by(User.name, ShiftAssignment.date)
    )
    assignments = assign_res.all()
    return sched, period, assignments


@router.get("/{schedule_id}/excel")
async def export_excel(
    schedule_id: str,
    manager: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db),
):
    sched, period, assignments = await _load_schedule_data(schedule_id, manager.tenant_id, db)
    buf = _build_excel(sched, period, assignments)
    filename = f"schedule_{period.label.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{schedule_id}/pdf")
async def export_pdf(
    schedule_id: str,
    manager: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db),
):
    sched, period, assignments = await _load_schedule_data(schedule_id, manager.tenant_id, db)
    buf = _build_pdf(sched, period, assignments)
    filename = f"schedule_{period.label.replace(' ', '_')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_excel(sched, period, assignments) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import timedelta

    wb = Workbook()
    ws = wb.active
    ws.title = "График"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    # Build date list
    start = period.start_date
    end = period.end_date
    days = []
    d = start
    while d <= end:
        days.append(d)
        d += timedelta(days=1)

    # Build staff → date → (shift_code, color, label) lookup
    data: dict[str, dict] = {}
    staff_order: list[tuple[str, str]] = []
    seen = set()
    for a, sp, u, st in assignments:
        if u.id not in seen:
            staff_order.append((u.id, u.name))
            seen.add(u.id)
        data.setdefault(u.id, {})[a.date] = (st.code, st.color_hex, f"{st.start_hour:02d}:{st.start_min:02d}–{st.end_hour:02d}:{st.end_min:02d}", st.duration_h)

    C_HEADER = "1A2E44"
    C_SUBHEAD = "2C4A63"

    def fill(hex_color):
        return PatternFill(fill_type="solid", fgColor=hex_color)

    def fnt(bold=False, size=9, color="000000"):
        return Font(bold=bold, size=size, color=color, name="Calibri")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 1: title
    total_cols = 1 + len(days) + 3
    end_col = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{end_col}1")
    c = ws["A1"]
    c.value = f"РАБОТЕН ГРАФИК  ·  {period.label.upper()}"
    c.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    c.fill = fill(C_HEADER)
    c.alignment = center
    ws.row_dimensions[1].height = 26

    # Row 2: day headers
    ws.merge_cells("A2:A3")
    ws["A2"].value = "Служител"
    ws["A2"].font = fnt(bold=True, size=10, color="FFFFFF")
    ws["A2"].fill = fill(C_HEADER)
    ws["A2"].alignment = center
    ws.column_dimensions["A"].width = 14

    DAYS_BG = ["Пон", "Вт", "Ср", "Чет", "Пет", "Съб", "Нед"]
    for i, day in enumerate(days):
        col = i + 2
        c2 = ws.cell(row=2, column=col)
        c2.value = DAYS_BG[day.weekday()]
        c2.font = fnt(bold=True, size=7, color="FFFFFF")
        c2.fill = fill(C_SUBHEAD)
        c2.alignment = center
        c3 = ws.cell(row=3, column=col)
        c3.value = day.day
        c3.font = fnt(size=7, color="888888")
        c3.alignment = center
        ws.column_dimensions[get_column_letter(col)].width = 7

    # Summary headers
    for j, lbl in enumerate(["8ч", "10ч", "12ч"]):
        col = len(days) + 2 + j
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        c = ws.cell(row=2, column=col)
        c.value = lbl
        c.font = fnt(bold=True, size=8)
        c.alignment = center
        ws.column_dimensions[get_column_letter(col)].width = 6

    # Data rows
    for bi, (uid, name) in enumerate(staff_order):
        row = bi + 4
        ws.row_dimensions[row].height = 22
        nc = ws.cell(row=row, column=1)
        nc.value = name
        nc.font = fnt(bold=True, size=10)
        nc.fill = fill("E8EEF5")
        nc.alignment = center

        cnt = {8: 0, 10: 0, 12: 0}
        for i, day in enumerate(days):
            col = i + 2
            entry = data.get(uid, {}).get(day)
            cell = ws.cell(row=row, column=col)
            if entry:
                code, color, label, dur = entry
                cell.value = label
                cell.fill = fill(color.lstrip("#"))
                cnt[dur] = cnt.get(dur, 0) + 1
            cell.font = fnt(size=6.5)
            cell.alignment = center

        for j, h in enumerate([8, 10, 12]):
            col = len(days) + 2 + j
            c = ws.cell(row=row, column=col)
            c.value = cnt.get(h, 0)
            c.font = fnt(bold=True, size=10)
            c.alignment = center

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_pdf(sched, period, assignments) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from datetime import timedelta
    import os

    font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    font_bold_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont("DV", font_path))
        pdfmetrics.registerFont(TTFont("DV-B", font_bold_path))
        FN, FB = "DV", "DV-B"
    else:
        FN, FB = "Helvetica", "Helvetica-Bold"

    start = period.start_date
    end = period.end_date
    days = []
    d = start
    while d <= end:
        days.append(d)
        d += timedelta(days=1)

    data_map: dict[str, dict] = {}
    staff_order: list[tuple[str, str]] = []
    seen: set[str] = set()
    for a, sp, u, st in assignments:
        if u.id not in seen:
            staff_order.append((u.id, u.name))
            seen.add(u.id)
        data_map.setdefault(u.id, {})[a.date] = (f"{st.start_hour:02d}-{st.end_hour:02d}", st.color_hex, st.duration_h)

    DAYS_BG = ["Пон", "Вт", "Ср", "Чет", "Пет", "Съб", "Нед"]

    hdr = ["Служител"] + [f"{DAYS_BG[d.weekday()]}\n{d.day}" for d in days] + ["8ч", "10ч", "12ч"]
    rows = [hdr]
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1A2E44")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), FB),
        ("FONTSIZE", (0, 0), (-1, 0), 6),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), FN),
        ("FONTSIZE", (0, 1), (-1, -1), 6),
        ("FONTNAME", (0, 1), (0, -1), FB),
        ("GRID", (0, 0), (-1, -1), 0.3, rl_colors.HexColor("#AAAAAA")),
        ("ROWHEIGHT", (0, 0), (-1, 0), 18),
        ("ROWHEIGHT", (0, 1), (-1, -1), 18),
    ]

    for bi, (uid, name) in enumerate(staff_order):
        ri = bi + 1
        row_data = [name]
        cnt = {8: 0, 10: 0, 12: 0}
        for di, day in enumerate(days):
            ci = di + 1
            entry = data_map.get(uid, {}).get(day)
            if entry:
                label, color, dur = entry
                row_data.append(label)
                style_cmds.append(("BACKGROUND", (ci, ri), (ci, ri), rl_colors.HexColor(color)))
                cnt[dur] = cnt.get(dur, 0) + 1
            else:
                row_data.append("")
        row_data += [str(cnt.get(8, 0)), str(cnt.get(10, 0)), str(cnt.get(12, 0))]
        rows.append(row_data)

    page = rl_landscape(A4)
    margin = 0.6 * cm
    usable_w = page[0] - 2 * margin
    name_w = 1.8 * cm
    sum_w = 0.65 * cm
    day_w = (usable_w - name_w - 3 * sum_w) / len(days)
    col_widths = [name_w] + [day_w] * len(days) + [sum_w] * 3

    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(style_cmds))

    title_style = ParagraphStyle("t", fontName=FB, fontSize=12,
                                 textColor=rl_colors.HexColor("#1A2E44"), alignment=1)
    elements = [
        Paragraph(f"РАБОТЕН ГРАФИК  ·  {period.label.upper()}", title_style),
        Spacer(1, 0.3 * cm),
        table,
    ]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=page,
                            rightMargin=margin, leftMargin=margin,
                            topMargin=0.7 * cm, bottomMargin=0.7 * cm)
    doc.build(elements)
    buf.seek(0)
    return buf
