"""
Renderer tests: xlsx and PDF outputs for the brief-#2 golden schedule open
cleanly, carry the client layout (week blocks, legend, quota columns,
Cyrillic) and include the verification sheet/section.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.render_pdf import render_pdf
from app.services.render_xlsx import render_xlsx
from packages.scheduler.orchestrator import generate
from tests.test_golden_briefs import make_spec_v2


@pytest.fixture(scope="module")
def golden():
    spec = make_spec_v2()
    return spec, generate(spec)


def test_xlsx_layout(golden):
    spec, report = golden
    data = render_xlsx(spec, report, month_label="юли 2026")
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["Месечен График", "Проверка"]

    ws = wb["Месечен График"]
    assert "МЕСЕЧЕН РАБОТЕН ГРАФИК" in ws.cell(row=1, column=1).value
    assert ws.cell(row=2, column=2).value == "СЕДМИЦА  1"
    assert ws.cell(row=3, column=2).value == "Пон"
    names = [ws.cell(row=r, column=1).value for r in range(5, 9)]
    assert names == ["Васил", "Ники", "Джедая", "Афродита"]
    # cells carry shift times and colored fills
    first = ws.cell(row=5, column=2)
    assert "–" in first.value
    assert first.fill.fgColor.rgb not in (None, "00000000")

    ver = wb["Проверка"]
    assert ver.cell(row=1, column=1).value == "Правило"
    rule_ids = {ver.cell(row=r, column=1).value for r in range(2, ver.max_row + 1)}
    assert "never_close_alone" in rule_ids
    assert "quota" in rule_ids


def test_xlsx_duration_summary_columns(golden):
    spec, report = golden
    wb = load_workbook(BytesIO(render_xlsx(spec, report)))
    ws = wb["Месечен График"]
    # 28 day columns → summaries start at col 30; durations are 8/10/12
    labels = [ws.cell(row=2, column=c).value for c in (30, 31, 32)]
    assert labels == ["8ч\nсмени", "10ч\nсмени", "12ч\nсмени"]
    # brief #1 quota preserved in brief #2: 4 weeks × (2×8h, 2×10h, 3×12h)
    for r in range(5, 9):
        counts = [ws.cell(row=r, column=c).value for c in (30, 31, 32)]
        assert counts == [8, 8, 12], counts


def test_pdf_renders_with_cyrillic(golden):
    spec, report = golden
    data = render_pdf(spec, report, month_label="юли 2026")
    assert data[:5] == b"%PDF-"
    assert len(data) > 10_000
    # embedded DejaVu subset present (Cyrillic-capable)
    assert b"DejaVu" in data
